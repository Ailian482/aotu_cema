# -*- coding:utf-8 -*-
# 作    者：Ailian
# 开发时间：2022/7/7 17:29

"""
Excel 单元格写入时的样式定义
"""

from openpyxl.styles import PatternFill, Font


def write_result(cell, row, column, status=1):
    """

    :param cell:
    :param row:
    :param column:
    :param status: 1 为Pass，其他为Failed
    :return:
    """
    # pass 样式
    if status == 1:
        # 写入内容
        cell(row=row, column=column).value = 'Pass'
        # 单元格样式定义：绿色+加粗
        cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')  # 定义单元格填充颜色
        cell(row=row, column=column).font = Font(bold=True)  # 定义字体加粗
    # failed 样式
    else:
        # 写入内容
        cell(row=row, column=column).value = 'Failed'
        # 单元格样式定义：红色+加粗
        cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')  # 定义单元格填充颜色
        cell(row=row, column=column).font = Font(bold=True)  # 定义字体加粗
